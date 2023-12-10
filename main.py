import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

def calculate():
    result = eval(entry.get())
    result_label.config(text="结果：" + str(result))
    history_text.insert(tk.END, entry.get() + "\n")
    entry.delete(0, tk.END)

def create_calculator_gui():
    global entry  # 将entry定义为全局变量
    window = tk.Tk()
    window.title("https://github.com/CUCPTT/Calculator")
    window.resizable(False, False)

    # 设置窗口图标
    window.iconbitmap("misc/favicon.ico")

    history_text = tk.Text(window, width=36, height=5, font=("Arial", 16), bg="white", relief="flat")
    history_text.grid(row=0, column=0, columnspan=6)

    entry = tk.Entry(window, width=36, font=("Arial", 16), relief="flat")
    entry.grid(row=1, column=0, columnspan=6)

    # 加载Excel文件
    workbook = load_workbook("misc/GUI.xlsx")
    sheet = workbook.active
    # 设置按钮样式
    button_style = ttk.Style()
    button_style.configure("Custom.TButton",
                           width = 6,
                           font=("Arial", 16),
                           bordercolor="#00008B",
                           relief=tk.FLAT,
                           borderwidth=0,
                           borderradius=0)

    # 生成按钮
    row_count = 2
    for row in sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=5):
        for cell in row:
            button_text = cell.value

            if button_text == "Back":
                # 创建删除按钮
                button = ttk.Button(window, text=button_text, command=lambda: entry.delete(len(entry.get())-1), style="Custom.TButton")
            else:
                # 创建普通按钮
                button = ttk.Button(window, text=button_text, command=lambda txt=button_text: entry.insert(tk.END, txt), style="Custom.TButton")
            
            button.grid(row=row_count, column=(cell.column - 1))

            # 判断是否需要换行
            if cell.column == 5:
                row_count += 1

    window.mainloop()

if __name__ == "__main__":
    create_calculator_gui()